"""
Excel Merger (Excel表格合并工具)
=================================
A desktop application to merge multiple Excel files into one workbook.
Each source file becomes a separate worksheet.

Features:
- Drag & drop Excel files into the window
- Full format preservation (styles, tables, merged cells, conditional formatting)
- Optional Excel COM automation for perfect fidelity
- .xls and .xlsx support
- Left sidebar navigation with extensible module slots
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import tkinter.font as tkfont
import os
import copy
import ctypes
from ctypes import wintypes

# =============================================================================
# Windows Drag & Drop (WM_DROPFILES) via ctypes
# =============================================================================

def _setup_wm_dropfiles(hwnd, on_files_dropped):
    """Install a WM_DROPFILES handler on a tkinter window.
    Returns a tuple of objects that must be kept alive."""
    try:
        user32 = ctypes.windll.user32
        shell32 = ctypes.windll.shell32
    except (AttributeError, OSError):
        return None

    GWL_WNDPROC = -4
    WM_DROPFILES = 0x0233

    ptr_size = ctypes.sizeof(ctypes.c_void_p)
    if ptr_size == 8:
        _GetWindowLong = user32.GetWindowLongPtrW
        _SetWindowLong = user32.SetWindowLongPtrW
        _LRESULT = ctypes.c_int64
    else:
        _GetWindowLong = user32.GetWindowLongW
        _SetWindowLong = user32.SetWindowLongW
        _LRESULT = ctypes.c_long

    _GetWindowLong.restype = ctypes.c_void_p
    _GetWindowLong.argtypes = [wintypes.HWND, ctypes.c_int]
    _SetWindowLong.restype = ctypes.c_void_p
    _SetWindowLong.argtypes = [wintypes.HWND, ctypes.c_int, ctypes.c_void_p]

    _CallWindowProc = user32.CallWindowProcW
    _CallWindowProc.restype = _LRESULT
    _CallWindowProc.argtypes = [ctypes.c_void_p, wintypes.HWND, ctypes.c_uint,
                                wintypes.WPARAM, wintypes.LPARAM]

    WNDPROC = ctypes.WINFUNCTYPE(_LRESULT, wintypes.HWND, ctypes.c_uint,
                                  wintypes.WPARAM, wintypes.LPARAM)

    original_wndproc = _GetWindowLong(hwnd, GWL_WNDPROC)

    @WNDPROC
    def new_wndproc(hwnd_inner, msg, wparam, lparam):
        if msg == WM_DROPFILES:
            hdrop = wparam
            count = shell32.DragQueryFileW(hdrop, 0xFFFFFFFF, None, 0)
            files = []
            for i in range(count):
                length = shell32.DragQueryFileW(hdrop, i, None, 0)
                buf = ctypes.create_unicode_buffer(length + 1)
                shell32.DragQueryFileW(hdrop, i, buf, length + 1)
                files.append(buf.value)
            shell32.DragFinish(hdrop)
            on_files_dropped(files)
            return 0
        return _CallWindowProc(original_wndproc, hwnd_inner, msg, wparam, lparam)

    _SetWindowLong(hwnd, GWL_WNDPROC, ctypes.cast(new_wndproc, ctypes.c_void_p))
    shell32.DragAcceptFiles(hwnd, True)
    return (new_wndproc, original_wndproc)


# =============================================================================
# Font Detection
# =============================================================================

def detect_available_font(root, preferred, default="TkDefaultFont"):
    """Try fonts in order, returning the first one available on the system."""
    available = {f.lower() for f in tkfont.families(root)}
    for font_name in preferred:
        if font_name.lower() in available:
            return font_name
    return default


# =============================================================================
# Colour Utilities
# =============================================================================

def hex_to_rgb(hx):
    hx = hx.lstrip('#')
    return tuple(int(hx[i:i+2], 16) for i in (0, 2, 4))

def rgb_to_hex(r, g, b):
    return f"#{int(r):02x}{int(g):02x}{int(b):02x}"

def darken_hex(hex_color, factor=0.85):
    r, g, b = hex_to_rgb(hex_color)
    return rgb_to_hex(int(r * factor), int(g * factor), int(b * factor))


# =============================================================================
# RoundedButton -- Canvas-based rounded button with gradient
# =============================================================================

class RoundedButton(tk.Canvas):
    """A rounded-rectangle button with gradient background and hover darkening."""

    def __init__(self, parent, text, command, *,
                 width=120, height=36, radius=8,
                 bg_color="#1a73e8", fg_color="white",
                 gradient_start=None, gradient_end=None,
                 font=None, state="normal", **kwargs):
        super().__init__(parent, width=width, height=height,
                         highlightthickness=0, borderwidth=0,
                         bg=parent.cget("bg") if hasattr(parent, "cget") else "#f0f4f8",
                         **kwargs)
        self._width = width
        self._height = height
        self._radius = radius
        self._bg_color = bg_color
        self._fg_color = fg_color
        self._gradient_start = gradient_start or darken_hex(bg_color, 0.78)
        self._gradient_end = gradient_end or bg_color
        self._text = text
        self._font = font or ("Microsoft YaHei UI", 10)
        self._command = command
        self._state = state

        self._draw_normal()
        self.bind("<Button-1>", self._on_click)
        self.bind("<Enter>", self._on_enter)
        self.bind("<Leave>", self._on_leave)

    # -- Drawing helpers --

    def _clip_to_rounded(self, y):
        """Return (x_start, x_end) for row y in a rounded rect."""
        w, h, r = self._width, self._height, self._radius
        if y < r:
            dx = int((r ** 2 - (r - y) ** 2) ** 0.5)
            return (dx, w - dx)
        elif y > h - r:
            dy = y - (h - r)
            dx = int((r ** 2 - (r - dy) ** 2) ** 0.5)
            return (dx, w - dx)
        return (0, w)

    def _draw_rounded_rect(self, tag, color_start, color_end):
        """Draw a rounded rectangle filled with vertical gradient."""
        self.delete(tag)
        w, h = self._width, self._height
        step = 2
        for y in range(0, h, step):
            t = y / max(h - 1, 1)
            r_val = int(hex_to_rgb(color_start)[0] + (hex_to_rgb(color_end)[0] - hex_to_rgb(color_start)[0]) * t)
            g_val = int(hex_to_rgb(color_start)[1] + (hex_to_rgb(color_end)[1] - hex_to_rgb(color_start)[1]) * t)
            b_val = int(hex_to_rgb(color_start)[2] + (hex_to_rgb(color_end)[2] - hex_to_rgb(color_start)[2]) * t)
            color = rgb_to_hex(r_val, g_val, b_val)
            x0, x1 = self._clip_to_rounded(y)
            self.create_rectangle(x0, y, x1, min(y + step, h),
                                  fill=color, outline="", tags=tag)

    def _overlay_darken(self, tag, factor=0.85):
        """Overlay a semi-transparent dark layer (by redrawing with darkened colors)."""
        # Simulate darkening by redrawing gradient with factored colors
        self.delete(tag)
        w, h = self._width, self._height
        step = 2
        for y in range(0, h, step):
            t = y / max(h - 1, 1)
            r1 = int(hex_to_rgb(self._gradient_start)[0] * factor)
            g1 = int(hex_to_rgb(self._gradient_start)[1] * factor)
            b1 = int(hex_to_rgb(self._gradient_start)[2] * factor)
            r2 = int(hex_to_rgb(self._gradient_end)[0] * factor)
            g2 = int(hex_to_rgb(self._gradient_end)[1] * factor)
            b2 = int(hex_to_rgb(self._gradient_end)[2] * factor)
            r_val = int(r1 + (r2 - r1) * t)
            g_val = int(g1 + (g2 - g1) * t)
            b_val = int(b1 + (b2 - b1) * t)
            color = rgb_to_hex(r_val, g_val, b_val)
            x0, x1 = self._clip_to_rounded(y)
            self.create_rectangle(x0, y, x1, min(y + step, h),
                                  fill=color, outline="", tags=tag)

    def _draw_text(self, color=None):
        self.delete("text")
        self.create_text(self._width // 2, self._height // 2,
                         text=self._text,
                         fill=color or self._fg_color,
                         font=self._font, tags="text")

    def _draw_normal(self):
        """Draw the normal state (gradient background)."""
        if self._state == "disabled":
            self._draw_rounded_rect("bg", "#b0b8c0", "#c8d0d8")
            self._draw_text("#e0e4e8")
        else:
            self._draw_rounded_rect("bg", self._gradient_start, self._gradient_end)
            self._draw_text(self._fg_color)

    # -- Event handlers --

    def _on_enter(self, event):
        if self._state == "normal":
            self._overlay_darken("bg", 0.85)
            self._draw_text(self._fg_color)

    def _on_leave(self, event):
        if self._state == "normal":
            self._draw_rounded_rect("bg", self._gradient_start, self._gradient_end)
            self._draw_text(self._fg_color)

    def _on_click(self, event):
        if self._state == "normal" and self._command:
            self._command()

    # -- Public API --

    def set_text(self, text):
        self._text = text
        self.delete("text")
        self.create_text(self._width // 2, self._height // 2,
                         text=text, fill=self._fg_color,
                         font=self._font, tags="text")

    def set_state(self, enabled):
        prev = self._state
        self._state = "normal" if enabled else "disabled"
        if prev != self._state:
            self._draw_normal()

    def set_colors(self, bg_color, gradient_start=None, gradient_end=None):
        self._bg_color = bg_color
        self._gradient_start = gradient_start or darken_hex(bg_color, 0.78)
        self._gradient_end = gradient_end or bg_color
        if self._state == "normal":
            self._draw_normal()


# =============================================================================
# Sidebar -- Left navigation panel
# =============================================================================

class Sidebar(tk.Frame):
    """Left vertical navigation panel with module selection."""

    def __init__(self, parent, *, width=200, bg_color="#2c3e50",
                 accent_color="#1a73e8", font=None, **kwargs):
        super().__init__(parent, width=width, bg=bg_color, **kwargs)
        self._width = width
        self._bg = bg_color
        self._accent = accent_color
        self._font = font or ("Microsoft YaHei UI", 10)
        self._font_title = (self._font[0], 14, "bold")
        self._font_version = (self._font[0], 8)
        self._items = {}       # module_id -> (frame, label)

        self.pack_propagate(False)

        # Title area
        title_frame = tk.Frame(self, bg=bg_color, height=60)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)
        tk.Label(title_frame, text="Excel 工具箱", font=self._font_title,
                 bg=bg_color, fg="white", anchor=tk.W, padx=18, pady=16
                 ).pack(fill=tk.X)

        # Separator
        tk.Frame(self, bg="#3d5068", height=1).pack(fill=tk.X, padx=12)

        # Nav items container
        self._nav_frame = tk.Frame(self, bg=bg_color)
        self._nav_frame.pack(fill=tk.BOTH, expand=True, pady=8)

        # Spacer
        spacer = tk.Frame(self, bg=bg_color)
        spacer.pack(fill=tk.BOTH, expand=True)

        # Version info
        tk.Frame(self, bg="#3d5068", height=1).pack(fill=tk.X, padx=12)
        tk.Label(self, text="v1.0", font=self._font_version,
                 bg=bg_color, fg="#8899aa", pady=6).pack()

    def add_nav_item(self, text, module_id, command, disabled=False):
        """Add a clickable navigation item. Returns the item frame."""
        item = tk.Frame(self._nav_frame, bg=self._bg, height=42, cursor="hand2")
        item.pack(fill=tk.X, padx=8, pady=1)
        item.pack_propagate(False)

        # Indicator bar (left accent line when active)
        indicator = tk.Frame(item, bg=self._bg, width=3)
        indicator.place(x=0, y=6, height=30)

        lbl = tk.Label(item, text=text, font=self._font,
                       bg=self._bg, fg="#8899aa" if disabled else "#c8d4e0",
                       anchor=tk.W, padx=14, pady=8)
        lbl.pack(fill=tk.X)

        if not disabled:
            def on_enter(e, f=item, l=lbl):
                f.config(bg="#34495e")
                l.config(bg="#34495e")

            def on_leave(e, f=item, l=lbl):
                f.config(bg=self._bg)
                l.config(bg=self._bg)

            def on_click(e, cmd=command):
                cmd()

            for w in (item, lbl):
                w.bind("<Enter>", on_enter)
                w.bind("<Leave>", on_leave)
                w.bind("<Button-1>", on_click)

        self._items[module_id] = (item, lbl, indicator, disabled)
        return item

    def set_active(self, module_id):
        """Highlight the given module as active."""
        for mid, (item, lbl, indicator, disabled) in self._items.items():
            if disabled:
                continue
            if mid == module_id:
                item.config(bg="#34495e")
                lbl.config(bg="#34495e", fg="white")
                indicator.config(bg=self._accent)
            else:
                item.config(bg=self._bg)
                lbl.config(bg=self._bg, fg="#c8d4e0")
                indicator.config(bg=self._bg)


# =============================================================================
# Excel Merging Logic
# =============================================================================

# Detect win32com availability once at module load
HAS_WIN32COM = False
try:
    import pythoncom
    import win32com.client as _win32
    _test = _win32.dynamic.Dispatch('Excel.Application')
    _test.Quit()
    del _test
    HAS_WIN32COM = True
except Exception:
    pass


class ExcelMerger:
    """Handles merging of Excel files with full format preservation."""

    _FORBIDDEN = str.maketrans({'\\': '_', '/': '_', '*': '_', '?': '_',
                                ':': '_', '[': '_', ']': '_'})

    # ---- Sheet naming ----

    @staticmethod
    def sanitize_sheet_name(name, max_len=31):
        name = name.translate(ExcelMerger._FORBIDDEN).strip().strip("'")
        return (name or "Sheet")[:max_len]

    @staticmethod
    def _unique_sheet_name(used, desired):
        base = ExcelMerger.sanitize_sheet_name(desired, 28)
        if base not in used:
            used[base] = 1
            return base
        used[base] += 1
        return f"{base}_{used[base]}"[:31]

    # ---- File validation ----

    @staticmethod
    def validate_file(path):
        """Validate an Excel file. Returns (is_valid, error_message)."""
        if not os.path.exists(path):
            return False, f"文件不存在: {path}"
        if os.path.getsize(path) == 0:
            return False, f"文件为空: {os.path.basename(path)}"
        ext = os.path.splitext(path)[1].lower()
        if ext == '.xlsb':
            return False, f"不支持 .xlsb 格式，请先另存为 .xlsx:\n{os.path.basename(path)}"
        if ext not in ('.xlsx', '.xlsm', '.xls'):
            return False, f"不支持的文件格式: {os.path.basename(path)}"
        try:
            if ext == '.xls':
                import xlrd
                wb = xlrd.open_workbook(path)
                wb.sheet_names()
            else:
                import openpyxl
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                wb.sheetnames
                wb.close()
        except Exception as e:
            return False, f"文件无法读取:\n{os.path.basename(path)}\n{str(e)[:120]}"
        return True, ""

    # ---- .xlsx full-format copy ----

    @staticmethod
    def _copy_cell_style(src_cell, dst_cell):
        """Deep-copy all style attributes from src to dst cell."""
        dst_cell.font = copy.copy(src_cell.font)
        dst_cell.fill = copy.copy(src_cell.fill)
        dst_cell.border = copy.copy(src_cell.border)
        dst_cell.alignment = copy.copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format
        dst_cell.protection = copy.copy(src_cell.protection)

    @staticmethod
    def _copy_tables(src_ws, dst_ws):
        """Copy table (ListObject) definitions."""
        try:
            for table in src_ws._tables:
                dst_ws.add_table(copy.copy(table))
        except Exception:
            pass  # Tables can be finicky across workbooks

    @staticmethod
    def _copy_conditional_formatting(src_ws, dst_ws):
        """Copy conditional formatting rules."""
        try:
            for cf in src_ws.conditional_formatting:
                dst_ws.conditional_formatting.add(copy.copy(cf))
        except Exception:
            pass

    @staticmethod
    def _copy_data_validations(src_ws, dst_ws):
        """Copy data validation rules."""
        try:
            if src_ws.data_validations:
                for dv in src_ws.data_validations.dataValidation:
                    dst_ws.add_data_validation(copy.copy(dv))
        except Exception:
            pass

    @staticmethod
    def _copy_sheet_settings(src_ws, dst_ws):
        """Copy sheet-level properties and page setup."""
        try:
            if src_ws.sheet_properties:
                dst_ws.sheet_properties = copy.copy(src_ws.sheet_properties)
        except Exception:
            pass
        try:
            if src_ws.sheet_format:
                dst_ws.sheet_format = copy.copy(src_ws.sheet_format)
        except Exception:
            pass

    @staticmethod
    def copy_xlsx_sheet_preserve(src_path, dst_wb, sheet_name):
        """Full-format sheet copy with styles, tables, conditional formatting, etc."""
        import openpyxl

        src_wb = openpyxl.load_workbook(src_path)
        try:
            name_used = {}
            for src_ws in src_wb.worksheets:
                if src_ws.max_row == 0 and src_ws.max_column == 0:
                    continue

                if len(src_wb.worksheets) > 1:
                    sn = ExcelMerger._unique_sheet_name(name_used,
                                                        f"{sheet_name}_{src_ws.title}")
                else:
                    sn = ExcelMerger.sanitize_sheet_name(sheet_name)

                dst_ws = dst_wb.create_sheet(title=sn)

                # Cell values + full styles
                for row in src_ws.iter_rows(min_row=1, max_row=src_ws.max_row,
                                            max_col=src_ws.max_column):
                    for cell in row:
                        dst_cell = dst_ws.cell(row=cell.row, column=cell.column)
                        dst_cell.value = cell.value
                        ExcelMerger._copy_cell_style(cell, dst_cell)

                # Merged cells
                for merge_range in src_ws.merged_cells.ranges:
                    dst_ws.merge_cells(str(merge_range))

                # Column widths
                for col_letter, col_dim in src_ws.column_dimensions.items():
                    if col_dim.width:
                        dst_ws.column_dimensions[col_letter].width = col_dim.width

                # Row heights
                for row_idx, row_dim in src_ws.row_dimensions.items():
                    if row_dim.height:
                        dst_ws.row_dimensions[row_idx].height = row_dim.height

                # Advanced features
                ExcelMerger._copy_tables(src_ws, dst_ws)
                ExcelMerger._copy_conditional_formatting(src_ws, dst_ws)
                ExcelMerger._copy_data_validations(src_ws, dst_ws)
                ExcelMerger._copy_sheet_settings(src_ws, dst_ws)

        finally:
            src_wb.close()

    # ---- .xls copy (enhanced: merged cells + dimensions) ----

    @staticmethod
    def copy_xls_sheet(src_path, dst_wb, sheet_name):
        """Copy .xls content including merged cells and dimensions."""
        try:
            import xlrd
        except ImportError:
            raise ImportError("读取 .xls 文件需要 xlrd 库。\n请运行: pip install xlrd")

        import openpyxl

        wb = xlrd.open_workbook(src_path, formatting_info=True)
        name_used = {}
        for sheet_idx, sheet in enumerate(wb.sheets()):
            if sheet.nrows == 0 and sheet.ncols == 0:
                continue

            if len(wb.sheet_names()) > 1:
                sn = ExcelMerger._unique_sheet_name(name_used,
                                                    f"{sheet_name}_{sheet.name}")
            else:
                sn = ExcelMerger.sanitize_sheet_name(sheet_name)

            dst_ws = dst_wb.create_sheet(title=sn)

            # Values
            for r in range(sheet.nrows):
                row_data = [sheet.cell_value(r, c) for c in range(sheet.ncols)]
                dst_ws.append(row_data)

            # Merged cells (xlrd uses 0-based, exclusive upper bound)
            try:
                if hasattr(sheet, 'merged_cells') and sheet.merged_cells:
                    for rlo, rhi, clo, chi in sheet.merged_cells:
                        first = openpyxl.utils.get_column_letter(clo + 1) + str(rlo + 1)
                        last = openpyxl.utils.get_column_letter(chi) + str(rhi)
                        dst_ws.merge_cells(f"{first}:{last}")
            except Exception:
                pass

            # Column widths (xlrd colinfo_map: col_idx -> Colinfo)
            try:
                if hasattr(sheet, 'colinfo_map'):
                    for col_idx, col_info in sheet.colinfo_map.items():
                        if col_info.width:
                            col_letter = openpyxl.utils.get_column_letter(col_idx + 1)
                            dst_ws.column_dimensions[col_letter].width = col_info.width / 7
            except Exception:
                pass

            # Row heights
            try:
                if hasattr(sheet, 'rowinfo_map'):
                    for row_idx, row_info in sheet.rowinfo_map.items():
                        if row_info.height:
                            dst_ws.row_dimensions[row_idx + 1].height = row_info.height / 20
            except Exception:
                pass

    # ---- win32com perfect merge ----

    @staticmethod
    def _merge_via_win32com(file_paths, output_path, progress_callback=None):
        """Merge using Excel COM automation for 100% fidelity."""
        import pythoncom
        import win32com.client as win32

        pythoncom.CoInitialize()
        excel = win32.dynamic.Dispatch('Excel.Application')
        excel.DisplayAlerts = False
        excel.Visible = False

        try:
            dst_wb = excel.Workbooks.Add()
            total = len(file_paths)
            sheets_created = 0

            for i, fp in enumerate(file_paths):
                fname = os.path.splitext(os.path.basename(fp))[0]
                msg = f"处理中 ({i + 1}/{total}): {fname}"
                if progress_callback:
                    progress_callback(int((i / total) * 90), msg)

                src_wb = excel.Workbooks.Open(fp)
                try:
                    for src_ws in src_wb.Worksheets:
                        src_ws.Copy(Before=dst_wb.Worksheets(1))
                        sheets_created += 1
                finally:
                    src_wb.Close(SaveChanges=False)

            # Rename sheets: win32com copies preserve original names
            # If source has only 1 sheet, rename to source filename
            # (Complex multi-sheet renaming omitted for simplicity --
            #  win32com copies already have meaningful names)

            # Remove extraneous default sheets
            while dst_wb.Worksheets.Count > sheets_created:
                try:
                    dst_wb.Worksheets(dst_wb.Worksheets.Count).Delete()
                except Exception:
                    break

            if progress_callback:
                progress_callback(95, "正在保存文件...")

            # 51 = xlOpenXMLWorkbook (.xlsx)
            dst_wb.SaveAs(output_path, FileFormat=51)
            dst_wb.Close(SaveChanges=False)

            if progress_callback:
                progress_callback(100, f"完成！共 {sheets_created} 个工作表")

            return sheets_created
        finally:
            excel.Quit()
            pythoncom.CoUninitialize()

    # ---- Main merge entry ----

    @staticmethod
    def merge(file_paths, output_path, progress_callback=None, use_win32com=False):
        """Merge Excel files. use_win32com=True uses Excel COM (requires Excel installed)."""
        import openpyxl

        if use_win32com and HAS_WIN32COM:
            has_xlsb = any(fp.lower().endswith('.xlsb') for fp in file_paths)
            if not has_xlsb:
                return ExcelMerger._merge_via_win32com(file_paths, output_path, progress_callback)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        total = len(file_paths)
        warnings = []

        for i, fp in enumerate(file_paths):
            fname = os.path.splitext(os.path.basename(fp))[0]
            sheet_name = ExcelMerger.sanitize_sheet_name(fname)

            # Validate
            valid, err = ExcelMerger.validate_file(fp)
            if not valid:
                warnings.append(err)
                if progress_callback:
                    progress_callback(int((i / total) * 90), f"跳过: {os.path.basename(fp)}")
                continue

            msg = f"处理中 ({i + 1}/{total}): {fname}"
            if progress_callback:
                progress_callback(int((i / total) * 90), msg)

            ext = fp.lower()
            try:
                if ext.endswith('.xls') and not ext.endswith('.xlsx'):
                    ExcelMerger.copy_xls_sheet(fp, wb, sheet_name)
                else:
                    ExcelMerger.copy_xlsx_sheet_preserve(fp, wb, sheet_name)
            except Exception as exc:
                warnings.append(f"{os.path.basename(fp)}: {str(exc)[:120]}")

        if len(wb.sheetnames) == 0:
            wb.create_sheet(title="空工作表")

        if progress_callback:
            progress_callback(95, "正在保存文件...")

        wb.save(output_path)

        if warnings:
            warning_msg = "\n".join(warnings[:5])
            if len(warnings) > 5:
                warning_msg += f"\n...还有 {len(warnings) - 5} 条警告"
            messagebox.showwarning("合并警告", warning_msg)

        sheets_count = len(wb.sheetnames)
        if progress_callback:
            progress_callback(100, f"完成！共 {sheets_count} 个工作表")

        return sheets_count


# =============================================================================
# GUI Application
# =============================================================================

class ExcelMergerApp:
    """Main GUI application with sidebar navigation."""

    EXCEL_EXTS = ('.xlsx', '.xls', '.xlsm')
    BG_COLOR = "#f0f4f8"
    ACCENT_COLOR = "#1a73e8"
    SIDEBAR_BG = "#2c3e50"
    SIDEBAR_WIDTH = 200

    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Excel Merger - Excel表格合并工具")
        self.root.geometry("960x560")
        self.root.minsize(740, 420)
        self.root.configure(bg=self.BG_COLOR)

        self.files = []
        self._dnd_refs = None
        self._active_module = None

        self._detect_fonts()
        self._build_ui()
        # Drag-drop installed in run() via after()

    # -- Font detection --

    def _detect_fonts(self):
        """Detect best available CJK fonts on this system."""
        ui_font = detect_available_font(
            self.root,
            preferred=["Microsoft YaHei UI", "Microsoft YaHei", "SimHei"],
            default="TkDefaultFont"
        )
        mono_font = detect_available_font(
            self.root,
            preferred=["Source Han Sans SC", "Noto Sans SC",
                       "Microsoft YaHei", "SimHei"],
            default="TkDefaultFont"
        )
        self.FONT_UI       = (ui_font, 10)
        self.FONT_UI_BOLD  = (ui_font, 10, "bold")
        self.FONT_HINT     = (ui_font, 12)
        self.FONT_STATUS   = (ui_font, 9)
        self.FONT_LIST     = (mono_font, 11)
        self.FONT_BTN      = (ui_font, 10)
        self.FONT_BTN_BIG  = (ui_font, 11, "bold")
        self.FONT_HEADER   = (ui_font, 9, "bold")

    # -- Drag & Drop ----------------------------------------------------------

    def _install_drag_drop(self):
        """Hook WM_DROPFILES (called after mainloop starts)."""
        self.root.update_idletasks()
        hwnd = self.root.winfo_id()
        if hwnd:
            self._dnd_refs = _setup_wm_dropfiles(hwnd, self._on_files_dropped)

    def _on_files_dropped(self, filepaths):
        added = 0
        for fp in filepaths:
            fp_lower = fp.lower()
            if fp_lower.endswith(self.EXCEL_EXTS) and fp not in self.files:
                self.files.append(fp)
                self.listbox.insert(tk.END, f"  {os.path.basename(fp)}")
                added += 1
        if added:
            self._refresh_ui()
            self._flash_drop_feedback(f"已添加 {added} 个文件")

    def _flash_drop_feedback(self, text, duration_ms=2500):
        """Flash the drop zone border + text to confirm a successful drop."""
        self.drop_frame.config(highlightbackground=self.ACCENT_COLOR,
                               highlightthickness=2)
        self.drop_label.config(text=f"✅ {text}", fg="#0d8043")
        self.root.after(duration_ms, self._restore_drop_style)

    def _restore_drop_style(self):
        self.drop_frame.config(highlightbackground="#c4c9cf",
                               highlightthickness=1)
        self.drop_label.config(
            text="📁  将 Excel 文件拖拽到此窗口，或点击下方按钮选择文件",
            fg="#5a6268"
        )

    # -- UI Construction ------------------------------------------------------

    def _build_ui(self):
        """Two-panel layout: sidebar + content area."""

        # ---- Root container ----
        self.root_container = tk.Frame(self.root, bg=self.BG_COLOR)
        self.root_container.pack(fill=tk.BOTH, expand=True)

        # ---- LEFT SIDEBAR ----
        self.sidebar = Sidebar(
            self.root_container,
            width=self.SIDEBAR_WIDTH,
            bg_color=self.SIDEBAR_BG,
            accent_color=self.ACCENT_COLOR,
            font=self.FONT_UI,
        )
        self.sidebar.pack(side=tk.LEFT, fill=tk.Y)

        # ---- RIGHT CONTENT AREA ----
        self.content_area = tk.Frame(self.root_container, bg=self.BG_COLOR)
        self.content_area.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # ---- MODULES ----
        self._modules = {}
        self._build_excel_merger_module()

        # Populate sidebar nav
        self._populate_sidebar()

        # Show default
        self.show_module("excel_merger")

    def _build_excel_merger_module(self):
        """Build the Excel Merger module UI."""
        frame = tk.Frame(self.content_area, bg=self.BG_COLOR)
        self._modules["excel_merger"] = frame

        main = tk.Frame(frame, bg=self.BG_COLOR)
        main.pack(fill=tk.BOTH, expand=True, padx=20, pady=15)

        # --- Drop zone ---
        drop_outer = tk.Frame(main, bg=self.BG_COLOR)
        drop_outer.pack(fill=tk.X, pady=5)

        self.drop_frame = tk.Frame(
            drop_outer, bg="white",
            highlightbackground="#c4c9cf", highlightthickness=1,
        )
        self.drop_frame.pack(fill=tk.X, ipady=14)

        self.drop_label = tk.Label(
            self.drop_frame,
            text="📁  将 Excel 文件拖拽到此窗口，或点击下方按钮选择文件",
            font=self.FONT_HINT, bg="white", fg="#5a6268", anchor=tk.CENTER,
            padx=10, pady=8,
        )
        self.drop_label.pack(fill=tk.X)

        # --- File list ---
        list_outer = tk.Frame(main, bg=self.BG_COLOR)
        list_outer.pack(fill=tk.BOTH, expand=True, pady=(12, 8))

        list_frame = tk.Frame(
            list_outer, bg="white",
            highlightbackground="#c4c9cf", highlightthickness=1,
        )
        list_frame.pack(fill=tk.BOTH, expand=True)

        header = tk.Frame(list_frame, bg="#e8ecf0", height=32)
        header.pack(fill=tk.X)
        header.pack_propagate(False)
        tk.Label(header, text="  待合并文件列表", font=self.FONT_HEADER,
                 bg="#e8ecf0", fg="#3c4043", anchor=tk.W
                 ).pack(side=tk.LEFT, fill=tk.X, expand=True)

        sb = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        self.listbox = tk.Listbox(
            list_frame, selectmode=tk.EXTENDED, yscrollcommand=sb.set,
            font=self.FONT_LIST, bg="white", fg="#202124",
            activestyle="none", borderwidth=0, highlightthickness=0,
        )
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True,
                          padx=(8, 0), pady=(0, 8))
        sb.config(command=self.listbox.yview)

        # Context menu
        self.listbox_menu = tk.Menu(self.listbox, tearoff=0)
        self.listbox_menu.add_command(label="移除选中项", command=self._remove_selected)
        self.listbox_menu.add_command(label="清空全部", command=self._clear_all)
        self.listbox.bind("<Button-3>", self._on_right_click)

        # --- Button row ---
        btn_row = tk.Frame(main, bg=self.BG_COLOR)
        btn_row.pack(fill=tk.X, pady=(6, 0))

        left = tk.Frame(btn_row, bg=self.BG_COLOR)
        left.pack(side=tk.LEFT)

        self.btn_add = self._make_rbtn(left, "➕  添加文件", self._browse_files,
                                       "#1a73e8", "#1557b0", "#1a73e8", 130, 36)
        self.btn_add.pack(side=tk.LEFT, padx=(0, 8))

        self.btn_remove = self._make_rbtn(left, "❌  移除选中", self._remove_selected,
                                          "#ea4335", "#c5221f", "#ea4335", 130, 36)
        self.btn_remove.pack(side=tk.LEFT, padx=(0, 8))

        self.btn_clear = self._make_rbtn(left, "🗑  清空列表", self._clear_all,
                                         "#5f6368", "#494c50", "#5f6368", 120, 36)
        self.btn_clear.pack(side=tk.LEFT)

        right = tk.Frame(btn_row, bg=self.BG_COLOR)
        right.pack(side=tk.RIGHT)

        self.btn_merge = self._make_rbtn(
            right, "🔀  合并生成", self._do_merge,
            "#1a73e8", "#1557b0", "#1a73e8", 155, 40,
        )
        self.btn_merge.pack(side=tk.RIGHT)

        # --- win32com checkbox ---
        checks_frame = tk.Frame(main, bg=self.BG_COLOR)
        checks_frame.pack(fill=tk.X, pady=(2, 0))
        self._use_win32com = tk.BooleanVar(value=HAS_WIN32COM)
        cb_text = "保留原始格式 (需要 Excel)" if HAS_WIN32COM else "保留原始格式 (未检测到 Excel)"
        self._cb_win32 = tk.Checkbutton(
            checks_frame, text=cb_text, variable=self._use_win32com,
            bg=self.BG_COLOR, font=self.FONT_STATUS,
            state=tk.NORMAL if HAS_WIN32COM else tk.DISABLED,
            activebackground=self.BG_COLOR,
        )
        self._cb_win32.pack(side=tk.LEFT)

        # --- Progress bar ---
        self.progress = ttk.Progressbar(main, mode='determinate')

        # --- Status bar ---
        self.status_var = tk.StringVar(value="就绪 — 请添加要合并的 Excel 文件")
        self.status_label = tk.Label(
            main, textvariable=self.status_var,
            anchor=tk.W, font=self.FONT_STATUS,
            bg="#e8ecf0", fg="#5a6268", padx=8, pady=4,
        )
        self.status_label.pack(fill=tk.X, pady=(10, 0))

    def _make_rbtn(self, parent, text, command, bg, gs, ge, w, h):
        """Create a RoundedButton with consistent styling."""
        return RoundedButton(
            parent, text=text, command=command,
            width=w, height=h, radius=8,
            bg_color=bg, fg_color="white",
            gradient_start=gs, gradient_end=ge,
            font=self.FONT_BTN,
        )

    def _populate_sidebar(self):
        """Add navigation items to the sidebar."""

        def nav_cmd(mod_id):
            return lambda: self.show_module(mod_id)

        self.sidebar.add_nav_item("📊  Excel 合并", "excel_merger",
                                  nav_cmd("excel_merger"))
        self.sidebar.add_nav_item("📋  待添加功能", "slot1",
                                  lambda: None, disabled=True)
        self.sidebar.add_nav_item("🔧  待添加功能", "slot2",
                                  lambda: None, disabled=True)
        self.sidebar.add_nav_item("⚙  待添加功能", "slot3",
                                  lambda: None, disabled=True)

    def show_module(self, module_id):
        """Switch visible module."""
        if self._active_module == module_id:
            return
        for frame in self._modules.values():
            frame.pack_forget()
        if module_id in self._modules:
            self._modules[module_id].pack(fill=tk.BOTH, expand=True)
        self._active_module = module_id
        self.sidebar.set_active(module_id)

    # -- Actions --------------------------------------------------------------

    def _browse_files(self):
        paths = filedialog.askopenfilenames(
            title="选择 Excel 文件",
            filetypes=[
                ("Excel 文件", "*.xlsx *.xls *.xlsm *.xlsb"),
                ("所有文件", "*.*"),
            ],
        )
        if paths:
            self._on_files_dropped(paths)

    def _on_right_click(self, event):
        try:
            idx = self.listbox.nearest(event.y)
            if not self.listbox.selection_includes(idx):
                self.listbox.selection_clear(0, tk.END)
                self.listbox.selection_set(idx)
            self.listbox_menu.tk_popup(event.x_root, event.y_root)
        finally:
            self.listbox_menu.grab_release()

    def _remove_selected(self):
        selected = list(self.listbox.curselection())
        if not selected:
            return
        for idx in reversed(selected):
            self.listbox.delete(idx)
            del self.files[idx]
        self._refresh_ui()

    def _clear_all(self):
        if not self.files:
            return
        self.files.clear()
        self.listbox.delete(0, tk.END)
        self._refresh_ui()

    def _refresh_ui(self):
        n = len(self.files)
        if n == 0:
            self.status_var.set("就绪 — 请添加要合并的 Excel 文件")
            self.btn_merge.set_state(False)
        else:
            self.status_var.set(f"共 {n} 个文件待合并")
            self.btn_merge.set_state(True)

    def _do_merge(self):
        if not self.files:
            messagebox.showinfo("提示", "请先添加要合并的 Excel 文件！")
            return

        output = filedialog.asksaveasfilename(
            title="保存合并后的文件",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx")],
            initialfile="合并结果.xlsx",
        )
        if not output:
            return

        self.progress.pack(fill=tk.X, pady=(0, 4), before=self.status_label)
        self.progress.config(value=0)
        self.root.update()

        def progress_cb(pct, msg):
            self.progress.config(value=pct)
            self.status_var.set(msg)
            self.root.update_idletasks()

        try:
            use_com = self._use_win32com.get() and HAS_WIN32COM
            sheet_count = ExcelMerger.merge(
                self.files, output,
                progress_callback=progress_cb,
                use_win32com=use_com,
            )
        except Exception as exc:
            self.progress.pack_forget()
            messagebox.showerror("合并出错", f"合并过程中发生错误：\n\n{exc}")
            self.status_var.set(f"错误: {exc}")
            return

        self.progress.pack_forget()
        self.status_var.set(f"✅ 合并完成！保存至: {os.path.basename(output)}")

        open_folder = messagebox.askyesno(
            "合并完成",
            f"成功合并 {len(self.files)} 个文件 → {sheet_count} 个工作表\n\n"
            f"保存位置：{output}\n\n"
            f"是否打开文件所在文件夹？"
        )
        if open_folder:
            os.startfile(os.path.dirname(os.path.abspath(output)))

    # -- Run ------------------------------------------------------------------

    def run(self):
        self._refresh_ui()

        # Center window
        self.root.update_idletasks()
        w, h = self.root.winfo_width(), self.root.winfo_height()
        sw = self.root.winfo_screenwidth()
        sh = self.root.winfo_screenheight()
        x = (sw - w) // 2
        y = (sh - h) // 2
        self.root.geometry(f"+{x}+{y}")

        # Install drag-drop after window is fully realized
        self.root.after(200, self._install_drag_drop)

        self.root.mainloop()


# =============================================================================
# Entry Point
# =============================================================================

def main():
    app = ExcelMergerApp()
    app.run()


if __name__ == '__main__':
    main()
